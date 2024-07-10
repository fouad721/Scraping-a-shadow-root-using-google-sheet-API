import asyncio
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime

# Initialize Google Sheets credentials
SCOPE = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']

# Ensure this path is correct relative to your script's location
CREDS_FILE = '**'  
SPREADSHEET_ID = '**' 

async def main():
    # Initialize Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Availability Data"
    sheet.append(["Date", "Availability"])

    # A list to keep track of all data before writing to Excel
    availability_data = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        last_response_time = None

        # Function to process the network response and update the last response time
        async def handle_response(response):
            nonlocal last_response_time
            if 'multi-room' in response.url:
                json_data = await response.json()
                for date_info in json_data['dates']:
                    date = date_info['date']
                    availability = "Available" if date_info['isAvailable'] else "Not Available"
                    # Add the date and availability to the list if not already present
                    if date not in map(lambda x: x[0], availability_data):
                        availability_data.append((date, availability))

                # Update the last response time
                last_response_time = asyncio.get_event_loop().time()

        page.on('response', handle_response)

        await page.goto('https://stay.ambergriscay.com/book/dates-of-stay?RoomTypeID=537387')
        await page.set_viewport_size({'width': 500, 'height': 768})

        # Continue scrolling and checking for new data until 5 seconds pass with no new information
        try:
            while True:
                # Scroll down
                await page.evaluate("window.scrollBy(0, 30000)")
                # Wait before checking for new responses
                await asyncio.sleep(2)

                # Check if 5 seconds have passed since the last new information
                if last_response_time and asyncio.get_event_loop().time() - last_response_time > 5:
                    print("No new information for 5 seconds, closing browser.")
                    break

        finally:
            # Sort the data by date from earliest to latest before writing to the Excel sheet
            availability_data.sort(key=lambda x: datetime.strptime(x[0], '%Y-%m-%d'))
            for date, availability in availability_data:
                sheet.append([date, availability])

            # Save the Excel workbook
            workbook.save("RoomAvailability.xlsx")
            print("Data has been written to RoomAvailability.xlsx and sorted by date.")

            # Close the browser
            await browser.close()

        # Now, read the saved Excel file and upload its contents to Google Sheets
        # Load the Excel workbook and get the data
        workbook = load_workbook(filename="RoomAvailability.xlsx")
        sheet = workbook.active
        excel_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            excel_data.append(row)
        workbook.close()

        # Initialize Google Sheets client
        credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDS_FILE, SCOPE)
        client = gspread.authorize(credentials)
        g_sheet = client.open_by_key(SPREADSHEET_ID).sheet1

        # Clear the Google Sheet and set the header
        g_sheet.clear()
        g_sheet.append_row(["Date", "Availability"])

        # Upload the data to Google Sheet starting from the second row
        g_sheet.update('A2', excel_data)
        print("Data has been uploaded to Google Sheet.")

# Run the main function
asyncio.run(main())
